"""Amend an existing harmonic_cohort_live run: backfill benchmark
``company_id`` FKs, re-evaluate, re-emit scoreboard + per_company JSON.

This lets us fix the "benchmark_rows_harmonic: 0" coverage bug in a
prior run without redoing the HTTP fetches.

Usage: ``python scripts/_harmonic_amend_eval.py <run_dir>``
"""

from __future__ import annotations

import json
import os
import sys
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]


def _bootstrap_env(run_dir: Path) -> None:
    db_path = (run_dir / "cohort.sqlite").resolve()
    cache_dir = (run_dir / "http_cache").resolve()
    run_artifact_dir = (run_dir / "run_artifacts").resolve()
    duckdb_path = (run_dir / "outputs" / "cohort.duckdb").resolve()
    os.environ["DB_URL"] = f"sqlite:///{db_path.as_posix()}"
    os.environ["CACHE_DIR"] = str(cache_dir)
    os.environ["RUN_ARTIFACT_DIR"] = str(run_artifact_dir)
    os.environ["DUCKDB_PATH"] = str(duckdb_path)


def _norm(s: str) -> str:
    out = s.lower().strip()
    for token in (",", ".", "(", ")", '"', "'"):
        out = out.replace(token, " ")
    for suffix in (
        " incorporated",
        " inc",
        " llc",
        " ltd",
        " limited",
        " corp",
        " corporation",
        " company",
        " co",
        " gmbh",
        " sa",
        " plc",
    ):
        if out.endswith(suffix):
            out = out[: -len(suffix)]
    return " ".join(out.split())


def main(argv: list[str]) -> int:
    if len(argv) != 1:
        print("usage: python scripts/_harmonic_amend_eval.py <run_dir>")
        return 2
    run_dir = Path(argv[0]).resolve()
    _bootstrap_env(run_dir)

    from sqlalchemy import select

    from headcount.db.engine import session_scope
    from headcount.db.enums import BenchmarkProvider
    from headcount.models.benchmark import BenchmarkObservation
    from headcount.models.company import Company
    from headcount.models.company_alias import CompanyAlias
    from headcount.models.headcount_estimate_monthly import HeadcountEstimateMonthly
    from headcount.resolution.resolver import _backfill_benchmark_links
    from headcount.review.evaluation import (
        EvaluationConfig,
        evaluate_against_benchmarks,
    )

    with session_scope() as session:
        _backfill_benchmark_links(session)

        targets_path = (
            REPO_ROOT
            / "test_source"
            / "Sample Employee Growth for High Priority Prospects.xlsx"
        )
        from openpyxl import load_workbook

        wb = load_workbook(targets_path, data_only=True, read_only=True)
        ws = wb["Harmonic April 8"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        idx = {h: i for i, h in enumerate(header)}
        targets = []
        for r in rows[1:]:
            name = r[idx["Company Name"]]
            if not name:
                continue
            targets.append(
                {
                    "name": str(name).strip(),
                    "headcount": r[idx["Headcount"]],
                    "growth_1y_pct": r[idx["Headcount % (365d)"]],
                    "growth_6m_pct": r[idx["Headcount % (180d)"]],
                }
            )

        all_companies = list(session.execute(select(Company)).scalars())
        lookup: dict[str, Any] = {}
        for c in all_companies:
            lookup[_norm(c.canonical_name)] = c
        for a in session.execute(select(CompanyAlias)).scalars():
            c = next((x for x in all_companies if x.id == a.company_id), None)
            if c is not None:
                lookup.setdefault(_norm(a.alias_name), c)

        resolved = []
        for t in targets:
            c = lookup.get(_norm(t["name"]))
            if c and c.id not in {x.id for x in resolved}:
                resolved.append(c)
        company_ids = [c.id for c in resolved]

        as_of = date(2026, 4, 1)
        scoreboard = evaluate_against_benchmarks(
            session,
            as_of_month=as_of,
            config=EvaluationConfig(),
            company_ids=company_ids,
            evaluated_at=datetime.now(tz=UTC),
        )

        per_company: list[dict[str, Any]] = []
        for t in targets:
            company = lookup.get(_norm(t["name"]))
            row: dict[str, Any] = {
                "harmonic_name": t["name"],
                "harmonic_headcount": t["headcount"],
                "harmonic_growth_1y_pct": t["growth_1y_pct"],
                "harmonic_growth_6m_pct": t["growth_6m_pct"],
                "company_id": company.id if company else None,
                "canonical_name": company.canonical_name if company else None,
            }
            if company is None:
                row["error"] = "no_canonical_match"
                per_company.append(row)
                continue
            est_rows = list(
                session.execute(
                    select(HeadcountEstimateMonthly)
                    .where(HeadcountEstimateMonthly.company_id == company.id)
                    .order_by(
                        HeadcountEstimateMonthly.created_at.desc(),
                        HeadcountEstimateMonthly.month,
                    )
                ).scalars()
            )
            if not est_rows:
                row["error"] = "no_estimate_rows"
                per_company.append(row)
                continue
            latest_version = est_rows[0].estimate_version_id
            by_month = {
                r.month: r for r in est_rows if r.estimate_version_id == latest_version
            }

            def _m(off: int, ref: date = as_of) -> date:
                y = ref.year + (ref.month - 1 - off) // 12
                mm = (ref.month - 1 - off) % 12 + 1
                return date(y, mm, 1)

            cur = by_month.get(as_of)
            t6 = by_month.get(_m(6))
            t12 = by_month.get(_m(12))
            row["estimate_current"] = cur.estimated_headcount if cur else None
            row["estimate_t_minus_6m"] = t6.estimated_headcount if t6 else None
            row["estimate_t_minus_1y"] = t12.estimated_headcount if t12 else None
            row["confidence_current"] = (
                round(cur.confidence_score, 3)
                if cur and cur.confidence_score is not None
                else None
            )
            if cur and t12 and t12.estimated_headcount:
                row["estimate_growth_1y_pct"] = round(
                    100.0 * (cur.estimated_headcount - t12.estimated_headcount)
                    / t12.estimated_headcount,
                    2,
                )
            if cur and t6 and t6.estimated_headcount:
                row["estimate_growth_6m_pct"] = round(
                    100.0 * (cur.estimated_headcount - t6.estimated_headcount)
                    / t6.estimated_headcount,
                    2,
                )
            bench_count = len(
                list(
                    session.execute(
                        select(BenchmarkObservation).where(
                            BenchmarkObservation.company_id == company.id,
                            BenchmarkObservation.provider == BenchmarkProvider.harmonic,
                        )
                    ).scalars()
                )
            )
            row["benchmark_rows_harmonic"] = bench_count
            per_company.append(row)

    (run_dir / "scoreboard.json").write_text(
        json.dumps(scoreboard.to_dict(), indent=2, default=str), encoding="utf-8"
    )
    (run_dir / "per_company.json").write_text(
        json.dumps(per_company, indent=2, default=str), encoding="utf-8"
    )
    print(f"amend_eval ok run_dir={run_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
