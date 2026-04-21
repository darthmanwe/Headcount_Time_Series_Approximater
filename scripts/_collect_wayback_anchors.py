"""Collect Wayback Machine historical anchors for the Harmonic cohort.

Targeted pass over the existing ``20260420T205059Z_postleak/cohort.sqlite``:
re-resolves the Harmonic subset, fans out the new ``WaybackObserver``, and
persists historical ``CompanyAnchorObservation`` rows so
``estimate_series`` can produce multi-anchor scaled-ratio estimates
instead of the degraded-current-only fallback.

The script is deliberately surgical: it does NOT re-run the live
``CompanyWebObserver`` / ``LinkedInPublicObserver`` (those already ran in
the source cohort and the cache is warm). It just layers the archival
signal on top.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from datetime import UTC, date, datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))


def _default_run_dir() -> Path:
    stamp = datetime.now(tz=UTC).strftime("%Y%m%dT%H%M%SZ")
    return REPO_ROOT / "data" / "runs" / "wayback_backfill" / stamp


def _bootstrap(run_dir: Path) -> None:
    """Point the process at the canonical DB + cache; per-run dir is
    only for local artifacts (logs / JSON summaries).

    Explicit ``DB_URL`` / ``CACHE_DIR`` in the environment still wins so
    an operator can sandbox an experimental retry when they want to.
    """

    run_artifact_dir = (run_dir / "run_artifacts").resolve()
    run_artifact_dir.mkdir(parents=True, exist_ok=True)

    if not os.environ.get("DB_URL", "").strip():
        canonical_db = (REPO_ROOT / "data" / "headcount.sqlite").resolve()
        canonical_db.parent.mkdir(parents=True, exist_ok=True)
        os.environ["DB_URL"] = f"sqlite:///{canonical_db.as_posix()}"
    if not os.environ.get("CACHE_DIR", "").strip():
        canonical_cache = (REPO_ROOT / "data" / "cache").resolve()
        canonical_cache.mkdir(parents=True, exist_ok=True)
        os.environ["CACHE_DIR"] = str(canonical_cache)
    os.environ["RUN_ARTIFACT_DIR"] = str(run_artifact_dir)


def _norm(s: str) -> str:
    out = s.lower().strip()
    for t in (",", ".", "(", ")", '"', "'"):
        out = out.replace(t, " ")
    for sfx in (
        " incorporated", " inc", " llc", " ltd", " limited", " corp",
        " corporation", " company", " co", " gmbh", " sa", " plc",
    ):
        if out.endswith(sfx):
            out = out[: -len(sfx)]
    return " ".join(out.split())


def _load_harmonic_names() -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(
        REPO_ROOT / "test_source" / "Sample Employee Growth for High Priority Prospects.xlsx",
        data_only=True, read_only=True,
    )
    ws = wb["Harmonic April 8"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = list(header).index("Company Name")
    return [str(r[idx]).strip() for r in rows[1:] if r[idx]]


async def _run(run_dir: Path) -> int:
    from headcount.config import get_settings
    from headcount.db.engine import session_scope
    from headcount.ingest.collect import (
        collect_anchors as _collect_anchors,
    )
    from headcount.ingest.collect import default_http_configs
    from headcount.ingest.http import FileCache, HttpClient
    from headcount.ingest.observers import WaybackObserver
    from headcount.models.company import Company
    from headcount.models.company_alias import CompanyAlias

    names = _load_harmonic_names()
    by_norm = {_norm(n): n for n in names}

    with session_scope() as session:
        cohort: dict[str, Company] = {}
        for c in session.query(Company).all():
            if _norm(c.canonical_name or "") in by_norm:
                cohort[c.id] = c
        for a in session.query(CompanyAlias).all():
            if _norm(a.alias_name or "") in by_norm and a.company_id not in cohort:
                c = session.get(Company, a.company_id)
                if c is not None:
                    cohort[c.id] = c
        companies = list(cohort.values())
        print(f"[wayback] cohort size: {len(companies)}")
        if not companies:
            print("[wayback] nothing to do")
            return 0

        settings = get_settings()
        cache = FileCache(settings.cache_dir)
        from headcount.ingest.raw_response_store import build_sink_from_session

        http = HttpClient(
            cache=cache,
            configs=default_http_configs(),
            transport=None,
            raw_response_sink=build_sink_from_session(session),
        )
        adapters: list[object] = [
            WaybackObserver(
                anchor_month=date(2026, 4, 1),
            )
        ]
        result = await _collect_anchors(
            session,
            adapters=adapters,  # type: ignore[arg-type]
            companies=companies,
            http_client=http,
            run_label=f"wayback_backfill:{run_dir.name}",
        )
        session.commit()
        summary = result.summary()
        print("[wayback] collect_anchors summary:")
        print(json.dumps(summary, indent=2, default=str))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "run_dir",
        type=Path,
        nargs="?",
        default=None,
        help=(
            "Directory for artifact output (logs, JSON). Defaults to"
            " data/runs/wayback_backfill/<ts>/. The database and HTTP"
            " cache are canonical - pass DB_URL / CACHE_DIR env vars to"
            " sandbox an experimental run."
        ),
    )
    args = parser.parse_args()
    run_dir = (args.run_dir or _default_run_dir()).resolve()
    run_dir.mkdir(parents=True, exist_ok=True)
    _bootstrap(run_dir)
    return asyncio.run(_run(run_dir))


if __name__ == "__main__":
    raise SystemExit(main())
