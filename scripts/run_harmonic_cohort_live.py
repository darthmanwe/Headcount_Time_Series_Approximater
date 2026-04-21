"""Blind live run against the Harmonic cohort (production-shape test).

This is the real-world capability test. The pipeline is fed ONLY the
data sources we will have in production:

- Seeded company records (name + domain + LinkedIn URL from the high-
  priority workbook).
- SEC EDGAR 10-K / 10-Q filings.
- Wikidata employee-count statements.
- Company website "about" / "careers" pages.
- Logged-out LinkedIn public company pages.
- LinkedIn OCR growth-trend charts (if tesseract is available).
- Manual anchor observer (no-op unless a file is supplied).

Harmonic and Zeeshan benchmark observations are deliberately loaded
*after* estimation finishes, purely as an evaluation reference. They
are never promoted to anchors, never seed the estimator, never inform
event parsing in this run. This keeps the test honest: we see what
our free-data pipeline actually produces, and we score it against the
Harmonic ground truth we'd never have in production.

Scope
-----
We run the pipeline over the ~24 companies that Harmonic has given us
ground-truth data for. Names are pulled from the Harmonic sheet and
resolved against the seeded ``Company``/``CompanyAlias`` tables. No
benchmark rows are consulted for scoping.

Artifacts (``data/runs/harmonic_live/<ts>/``)
---------------------------------------------
- ``pipeline.json``           - per-stage summary.
- ``scoreboard.json``         - evaluation harness output.
- ``per_company.json``        - Harmonic target vs our estimate, per company.
- ``production_coverage.json``- anchors produced per company, bucketed
                                by source. Shows what the free-data
                                pipeline actually saw.
- ``issues.json``             - structured errors / warnings.

Usage
-----

``python scripts/run_harmonic_cohort_live.py [--mode live-full|live-safe]``
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]


def _canonical_db_url() -> str:
    """Absolute ``sqlite:///`` URL for the long-lived, shared database.

    Every cohort run, retry, and backfill writes into one file so
    legitimate observations compound across runs. Callers that need
    true isolation can still override via ``DB_URL`` before launching
    the script.
    """

    env_override = os.environ.get("DB_URL", "").strip()
    if env_override:
        return env_override
    canonical = (REPO_ROOT / "data" / "headcount.sqlite").resolve()
    canonical.parent.mkdir(parents=True, exist_ok=True)
    return f"sqlite:///{canonical.as_posix()}"


def _canonical_cache_dir() -> Path:
    """Shared HTTP cache directory.

    Same reasoning as the DB: a successful fetch against LinkedIn or
    Wayback in one run should be reusable by the next one, so we point
    every run at the same cache tree unless ``CACHE_DIR`` is already
    set.
    """

    env_override = os.environ.get("CACHE_DIR", "").strip()
    if env_override:
        return Path(env_override)
    return (REPO_ROOT / "data" / "cache").resolve()


def _bootstrap_env(run_dir: Path) -> None:
    """Wire env vars for the canonical DB + cache, per-run artifact dirs.

    Previously this created per-run DB and cache files under ``run_dir``,
    which meant every cohort run started from zero - every successful
    LinkedIn / Wayback / company-web fetch vanished when the run
    finished. Now DB_URL and CACHE_DIR point at long-lived locations
    by default; only artifact dirs (logs, scoreboards, DuckDB export)
    stay per-run so experiments remain inspectable.
    """

    run_artifact_dir = (run_dir / "run_artifacts").resolve()
    duckdb_path = (run_dir / "outputs" / "cohort.duckdb").resolve()
    cache_dir = _canonical_cache_dir()
    for p in (cache_dir, run_artifact_dir, duckdb_path.parent):
        p.mkdir(parents=True, exist_ok=True)
    os.environ["DB_URL"] = _canonical_db_url()
    os.environ["CACHE_DIR"] = str(cache_dir)
    os.environ["RUN_ARTIFACT_DIR"] = str(run_artifact_dir)
    os.environ["DUCKDB_PATH"] = str(duckdb_path)


@dataclass
class Issue:
    stage: str
    severity: str  # "warn" | "error" | "gate"
    company: str | None
    message: str
    detail: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "stage": self.stage,
            "severity": self.severity,
            "company": self.company,
            "message": self.message,
            "detail": self.detail,
        }


@dataclass
class RunContext:
    run_dir: Path
    mode: str
    issues: list[Issue] = field(default_factory=list)

    def add(
        self,
        *,
        stage: str,
        severity: str,
        message: str,
        company: str | None = None,
        detail: dict[str, Any] | None = None,
    ) -> None:
        self.issues.append(
            Issue(
                stage=stage,
                severity=severity,
                company=company,
                message=message,
                detail=detail or {},
            )
        )


HARMONIC_SHEET = "Harmonic April 8"


def _load_harmonic_targets() -> list[dict[str, Any]]:
    """Read the 24 Harmonic target rows from the workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(
        REPO_ROOT / "test_source" / "Sample Employee Growth for High Priority Prospects.xlsx",
        data_only=True,
        read_only=True,
    )
    ws = wb[HARMONIC_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}

    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        name = r[idx["Company Name"]]
        if not name:
            continue
        out.append(
            {
                "name": str(name).strip(),
                "headcount": r[idx["Headcount"]],
                "growth_1y_pct": r[idx["Headcount % (365d)"]],
                "growth_6m_pct": r[idx["Headcount % (180d)"]],
                "growth_3m_pct": r[idx["Headcount % (90d)"]],
                "growth_1m_pct": r[idx["Headcount % (30d)"]],
            }
        )
    return out


def _run_alembic_upgrade() -> None:
    from alembic import command
    from alembic.config import Config

    cfg = Config(str(REPO_ROOT / "alembic.ini"))
    cfg.set_main_option("script_location", str(REPO_ROOT / "migrations"))
    cfg.set_main_option("sqlalchemy.url", os.environ["DB_URL"])
    command.upgrade(cfg, "head")


def _now_stamp() -> str:
    return datetime.now(tz=UTC).strftime("%Y%m%dT%H%M%SZ")


def _norm(s: str) -> str:
    """Loose normalized form for fuzzy name matching."""
    out = s.lower().strip()
    for token in (",", ".", "(", ")", '"', "'"):
        out = out.replace(token, " ")
    # Drop common legal suffixes so "Foo Inc" matches "Foo".
    for suffix in (
        " incorporated",
        " inc",
        " llc",
        " ltd",
        " limited",
        " corp",
        " corporation",
        " company",
        " co",
        " gmbh",
        " sa",
        " plc",
    ):
        if out.endswith(suffix):
            out = out[: -len(suffix)]
    return " ".join(out.split())


async def _collect_anchors_scoped(
    session: Any,
    companies: list[Any],
    *,
    mode: str,
    rate_guard: Any | None = None,
    run_label: str | None = None,
) -> dict[str, object]:
    """Dispatch the full public-source observer stack against a scoped company list.

    ``rate_guard`` is an optional shared :class:`LinkedInRateGuard`.
    When provided, the LinkedIn observer adopts it instead of building
    a private one, so slug-backfill and observer requests share a
    single budget / breaker / cooldown clock for the whole run.
    """
    from headcount.config import get_settings
    from headcount.ingest.collect import (
        collect_anchors as _collect_anchors,
    )
    from headcount.ingest.collect import (
        default_http_configs,
    )
    from headcount.ingest.http import FileCache, HttpClient
    from headcount.ingest.observers import (
        CompanyWebObserver,
        LinkedInPublicObserver,
        ManualAnchorObserver,
        SECObserver,
        WaybackObserver,
        WikidataObserver,
    )
    from headcount.ingest.raw_response_store import build_sink_from_session

    settings = get_settings()
    cache = FileCache(settings.cache_dir)
    # Plan C: every live response this cohort run sees is mirrored into
    # ``raw_response`` so a parser-version bump or a reparse-all pass
    # can reprocess the same bytes without re-hitting LinkedIn /
    # company websites / Wayback. Sink errors are swallowed inside the
    # sink itself - archival is best-effort and never fails a fetch.
    raw_sink = build_sink_from_session(session)
    http = HttpClient(
        cache=cache,
        configs=default_http_configs(),
        transport=None,  # live
        raw_response_sink=raw_sink,
    )
    adapters: list[object] = [ManualAnchorObserver(), SECObserver(), WikidataObserver()]
    if mode in {"live-safe", "live-full"}:
        adapters.append(CompanyWebObserver())
        # Wayback is included from live-safe onward because the only
        # outbound host is archive.org; it never touches LinkedIn or
        # target-company infra. Running it in both modes gives the
        # estimator real historical anchors to pair with the current-
        # month live anchor, which is what unlocks the 6m/1y/2y growth
        # metrics.
        adapters.append(WaybackObserver())
    if mode == "live-full":
        adapters.append(LinkedInPublicObserver(rate_guard=rate_guard))
    result = await _collect_anchors(
        session,
        adapters=adapters,  # type: ignore[arg-type]
        companies=companies,
        http_client=http,
        run_label=run_label,
    )
    return result.summary()


def _build_ocr_observer(ctx: RunContext) -> object | None:
    """Try to construct the OCR observer; on failure record a warn-level issue."""
    try:
        from headcount.ingest.observers.linkedin_ocr import (
            LinkedInGrowthTrendObserver,
        )

        return LinkedInGrowthTrendObserver()
    except Exception as exc:
        ctx.add(
            stage="collect_employment",
            severity="warn",
            message="OCR observer construction failed",
            detail={"error": str(exc)},
        )
        return None


def _score_harmonic_cohort(
    session: Any,
    *,
    as_of: date,
    company_ids: list[str],
) -> dict[str, Any]:
    from headcount.review.evaluation import (
        EvaluationConfig,
        evaluate_against_benchmarks,
    )

    scoreboard = evaluate_against_benchmarks(
        session,
        as_of_month=as_of,
        config=EvaluationConfig(),
        company_ids=company_ids,
        evaluated_at=datetime.now(tz=UTC),
    )
    return scoreboard.to_dict()


def _production_coverage(
    session: Any, *, company_ids: list[str]
) -> dict[str, Any]:
    """For every in-scope company, report what the free-data pipeline
    actually produced: anchor count bucketed by source_name."""
    from collections import Counter

    from sqlalchemy import select

    from headcount.models.company import Company
    from headcount.models.company_anchor_observation import CompanyAnchorObservation
    from headcount.models.source_observation import SourceObservation

    cov: dict[str, Any] = {"per_company": {}, "totals": Counter()}

    for cid in company_ids:
        company = session.execute(
            select(Company).where(Company.id == cid)
        ).scalar_one_or_none()
        if company is None:
            continue

        stmt = (
            select(CompanyAnchorObservation, SourceObservation)
            .outerjoin(
                SourceObservation,
                SourceObservation.id == CompanyAnchorObservation.source_observation_id,
            )
            .where(CompanyAnchorObservation.company_id == cid)
        )
        rows = list(session.execute(stmt).all())

        per_src: Counter[str] = Counter()
        samples: list[dict[str, Any]] = []
        for anchor, src in rows:
            source_name = src.source_name.value if src is not None else "unknown"
            per_src[source_name] += 1
            cov["totals"][source_name] += 1
            samples.append(
                {
                    "source": source_name,
                    "month": anchor.anchor_month.isoformat(),
                    "point": anchor.headcount_value_point,
                    "min": anchor.headcount_value_min,
                    "max": anchor.headcount_value_max,
                    "kind": anchor.headcount_value_kind.value,
                    "confidence": round(anchor.confidence, 3),
                }
            )

        cov["per_company"][company.canonical_name] = {
            "company_id": cid,
            "anchors_total": len(rows),
            "by_source": dict(per_src),
            "samples": samples[:10],
        }

    cov["totals"] = dict(cov["totals"])
    return cov


def _per_company_report(
    session: Any,
    *,
    harmonic_targets: list[dict[str, Any]],
    as_of: date,
    company_map: dict[str, Any],
) -> list[dict[str, Any]]:
    """Render one row per Harmonic target with our (blind) estimate side-by-side."""
    from sqlalchemy import select

    from headcount.db.enums import BenchmarkProvider
    from headcount.models.benchmark import BenchmarkObservation
    from headcount.models.headcount_estimate_monthly import HeadcountEstimateMonthly

    out: list[dict[str, Any]] = []
    for target in harmonic_targets:
        company = company_map.get(_norm(target["name"]))
        row: dict[str, Any] = {
            "harmonic_name": target["name"],
            "harmonic_headcount": target["headcount"],
            "harmonic_growth_1y_pct": target["growth_1y_pct"],
            "harmonic_growth_6m_pct": target["growth_6m_pct"],
            "company_id": company.id if company else None,
            "canonical_name": company.canonical_name if company else None,
            "estimate_current": None,
            "estimate_t_minus_6m": None,
            "estimate_t_minus_1y": None,
            "estimate_growth_1y_pct": None,
            "estimate_growth_6m_pct": None,
            "confidence_current": None,
            "error": None,
        }
        if company is None:
            row["error"] = "no_canonical_match"
            out.append(row)
            continue

        stmt = (
            select(HeadcountEstimateMonthly)
            .where(HeadcountEstimateMonthly.company_id == company.id)
            .order_by(
                HeadcountEstimateMonthly.created_at.desc(),
                HeadcountEstimateMonthly.month,
            )
        )
        rows = list(session.execute(stmt).scalars())
        if not rows:
            row["error"] = "no_estimate_rows"
            out.append(row)
            continue

        latest_version_id = rows[0].estimate_version_id
        by_month = {
            r.month: r for r in rows if r.estimate_version_id == latest_version_id
        }

        def _month(offset_months: int) -> date:
            year = as_of.year + (as_of.month - 1 - offset_months) // 12
            m_mod = (as_of.month - 1 - offset_months) % 12 + 1
            return date(year, m_mod, 1)

        current = by_month.get(as_of)
        t6 = by_month.get(_month(6))
        t12 = by_month.get(_month(12))
        row["estimate_current"] = current.estimated_headcount if current else None
        row["estimate_t_minus_6m"] = t6.estimated_headcount if t6 else None
        row["estimate_t_minus_1y"] = t12.estimated_headcount if t12 else None
        row["confidence_current"] = (
            round(current.confidence_score, 3)
            if current and current.confidence_score is not None
            else None
        )

        if current and t12 and t12.estimated_headcount:
            row["estimate_growth_1y_pct"] = round(
                100.0
                * (current.estimated_headcount - t12.estimated_headcount)
                / t12.estimated_headcount,
                2,
            )
        if current and t6 and t6.estimated_headcount:
            row["estimate_growth_6m_pct"] = round(
                100.0
                * (current.estimated_headcount - t6.estimated_headcount)
                / t6.estimated_headcount,
                2,
            )

        bench_stmt = select(BenchmarkObservation).where(
            BenchmarkObservation.company_id == company.id,
            BenchmarkObservation.provider == BenchmarkProvider.harmonic,
        )
        row["benchmark_rows_harmonic"] = len(list(session.execute(bench_stmt).scalars()))
        out.append(row)
    return out


def _resolve_harmonic_cohort(
    session: Any,
    *,
    harmonic_targets: list[dict[str, Any]],
) -> tuple[list[Any], dict[str, Any], list[str]]:
    """Match Harmonic target names to seeded Company rows via canonical
    name and aliases. Returns (companies, lookup, unmatched_names)."""
    from sqlalchemy import select

    from headcount.models.company import Company
    from headcount.models.company_alias import CompanyAlias

    all_companies = list(session.execute(select(Company)).scalars())
    lookup: dict[str, Any] = {}
    for c in all_companies:
        lookup[_norm(c.canonical_name)] = c
    alias_rows = list(session.execute(select(CompanyAlias)).scalars())
    company_by_id = {c.id: c for c in all_companies}
    for a in alias_rows:
        c = company_by_id.get(a.company_id)
        if c is not None:
            lookup.setdefault(_norm(a.alias_name), c)

    matched: list[Any] = []
    seen_ids: set[str] = set()
    unmatched: list[str] = []
    for t in harmonic_targets:
        c = lookup.get(_norm(t["name"]))
        if c is None:
            unmatched.append(t["name"])
            continue
        if c.id not in seen_ids:
            matched.append(c)
            seen_ids.add(c.id)
    return matched, lookup, unmatched


def _parse_cohort_slice(spec: str) -> tuple[int, int]:
    """Parse ``--cohort-slice N/M`` into validated integers.

    Raises :class:`ValueError` on any malformed or out-of-range input.
    Kept module-level so cohort sharding can be unit-tested without
    bringing the whole runner along.
    """

    try:
        shard_str, total_str = spec.split("/", 1)
        shard = int(shard_str)
        total = int(total_str)
    except (ValueError, AttributeError) as exc:
        raise ValueError(
            f"--cohort-slice must look like 'N/M', got {spec!r}"
        ) from exc
    if shard < 1 or total < 1 or shard > total:
        raise ValueError(f"--cohort-slice out of range: {spec!r}")
    return shard, total


def _apply_cohort_slice(
    companies: list[Any], spec: str
) -> tuple[list[Any], dict[str, Any]]:
    """Return the requested shard of ``companies`` plus a meta dict.

    The companies are sorted by canonical name (with id as tiebreaker)
    before slicing so the same ``N/M`` produces the same subset across
    runs - a hard requirement for multi-day production runs that need
    to resume on the same row of the workbook.
    """

    shard, total = _parse_cohort_slice(spec)
    ordered = sorted(companies, key=lambda c: (c.canonical_name or c.id))
    sliced = [c for i, c in enumerate(ordered) if (i % total) + 1 == shard]
    return sliced, {
        "shard": shard,
        "total": total,
        "in_shard": len(sliced),
        "skipped": len(ordered) - len(sliced),
    }


def _main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--mode",
        default="live-full",
        choices=["offline", "live-safe", "live-full"],
    )
    parser.add_argument(
        "--run-dir",
        default=None,
        help="Override output directory; defaults to data/runs/harmonic_live/<ts>/",
    )
    parser.add_argument(
        "--start", default="2020-01", help="First estimate month (YYYY-MM)."
    )
    parser.add_argument(
        "--as-of",
        default=None,
        help="Reference month for current/6m/1y (defaults to 2026-04).",
    )
    parser.add_argument(
        "--cohort-slice",
        default=None,
        help=(
            "Restrict the cohort to shard N of M (1-indexed, e.g. 1/4)."
            " Use to spread a 250-2000 company batch across multiple"
            " process runs or days while keeping the LinkedIn budget low."
        ),
    )
    parser.add_argument(
        "--retry-breaker-skips",
        action="store_true",
        help=(
            "After the first pass, sleep through the LinkedIn breaker"
            " cooldown and re-collect anchors only for companies parked"
            " by a circuit-open short-circuit. Recommended for"
            " unattended multi-hour runs; off by default for tight"
            " local feedback loops."
        ),
    )
    parser.add_argument(
        "--breaker-recovery-floor-seconds",
        type=float,
        default=0.0,
        help=(
            "Minimum cooldown (seconds) to sleep before the"
            " breaker-recovery pass, even if the breaker never tripped."
            " The script takes max(this floor, guard cooldown, 120s)."
            " Use higher values (eg 900) when LinkedIn has been issuing"
            " 999s without the breaker fully tripping."
        ),
    )
    parser.add_argument(
        "--run-label",
        default=None,
        help=(
            "Human-readable tag attached to the Run row, e.g."
            " 'harmonic_cohort:postleak' or 'wayback_backfill'. Defaults"
            " to 'harmonic_cohort:<run_dir_basename>' so each cohort run"
            " stays filterable in the shared canonical DB."
        ),
    )
    args = parser.parse_args(argv)

    run_dir = (
        Path(args.run_dir).resolve()
        if args.run_dir
        else REPO_ROOT / "data" / "runs" / "harmonic_live" / _now_stamp()
    )
    run_dir.mkdir(parents=True, exist_ok=True)

    run_label = args.run_label or f"harmonic_cohort:{run_dir.name}"

    _bootstrap_env(run_dir)
    _run_alembic_upgrade()

    from headcount.config import get_settings
    from headcount.db.engine import session_scope
    from headcount.db.enums import PriorityTier
    from headcount.estimate.pipeline import estimate_series as run_estimate
    from headcount.ingest.collect import default_http_configs
    from headcount.ingest.employment import collect_employment as _collect_emp
    from headcount.ingest.http import FileCache, HttpClient
    from headcount.ingest.seeds import import_candidates, load_benchmarks
    from headcount.parsers import merge_events, promote_benchmark_events
    from headcount.resolution import resolve_candidates
    from headcount.ingest.linkedin_guard import LinkedInRateGuard
    from headcount.resolution.linkedin_resolver import backfill_linkedin_slugs
    from headcount.resolution.resolver import _backfill_benchmark_links

    ctx = RunContext(run_dir=run_dir, mode=args.mode)
    summary: dict[str, Any] = {
        "mode": args.mode,
        "run_label": run_label,
        "stages": {},
        "run_dir": str(run_dir),
    }

    def _month(raw: str) -> date:
        raw = raw.strip()
        if len(raw) == 7:
            raw = f"{raw}-01"
        return date.fromisoformat(raw).replace(day=1)

    start = _month(args.start)
    as_of = _month(args.as_of) if args.as_of else date(2026, 4, 1)
    summary["as_of"] = as_of.isoformat()

    harmonic_targets = _load_harmonic_targets()
    summary["harmonic_targets"] = len(harmonic_targets)

    companies_workbook = (
        REPO_ROOT / "test_source" / "High Priority Companies_01.04.2026.xlsx"
    )
    benchmarks_workbook = (
        REPO_ROOT / "test_source" / "Sample Employee Growth for High Priority Prospects.xlsx"
    )

    with session_scope() as session:
        # --- Stage 1: seed companies (names only; NO benchmarks yet) ---
        t0 = time.time()
        seed = import_candidates(session, companies_workbook)
        summary["stages"]["seed_companies"] = {
            "rows_imported": seed.rows_imported,
            "rows_skipped": seed.rows_skipped,
            "elapsed_ms": int((time.time() - t0) * 1000),
        }

        # --- Stage 2: canonicalize ---
        t0 = time.time()
        canon = resolve_candidates(
            session, default_priority_tier=PriorityTier.P1, only_pending=True
        )
        summary["stages"]["canonicalize"] = {
            "candidates_scanned": canon.candidates_scanned,
            "companies_created": canon.companies_created,
            "aliases_created": canon.aliases_created,
            "elapsed_ms": int((time.time() - t0) * 1000),
        }

        # --- Stage 3: parse events (no benchmarks loaded = no-op) ---
        t0 = time.time()
        promote = promote_benchmark_events(session, only_pending=True)
        merged = merge_events(session)
        summary["stages"]["parse_events"] = {
            "promoted": promote.promoted,
            "groups_collapsed": merged.groups_collapsed,
            "elapsed_ms": int((time.time() - t0) * 1000),
        }

        # --- Stage 4: resolve Harmonic cohort by name against Company/alias ---
        t0 = time.time()
        companies, company_map, unmatched = _resolve_harmonic_cohort(
            session, harmonic_targets=harmonic_targets
        )
        for name in unmatched:
            ctx.add(
                stage="canonicalize",
                severity="warn",
                message="Harmonic target not resolved to a canonical company",
                company=name,
            )
        # --- Optional: shard the cohort ---
        # ``--cohort-slice 2/4`` keeps only the second of four equal
        # slices. The slice is taken AFTER deterministic name-sort by
        # ``_resolve_harmonic_cohort`` so the same N/M produces the
        # same set across runs.
        slice_meta: dict[str, Any] | None = None
        if args.cohort_slice:
            try:
                companies, slice_meta = _apply_cohort_slice(
                    companies, args.cohort_slice
                )
            except ValueError as exc:
                raise SystemExit(str(exc)) from exc

        summary["stages"]["pick_cohort"] = {
            "harmonic_targets": len(harmonic_targets),
            "resolved_companies": len(companies),
            "unmatched": unmatched,
            "elapsed_ms": int((time.time() - t0) * 1000),
        }
        if slice_meta is not None:
            summary["stages"]["pick_cohort"]["cohort_slice"] = slice_meta
        company_ids = [c.id for c in companies]

        # --- Stage 4.4: shared LinkedIn rate guard ---
        # One guard for the entire run so resolver + observer share a
        # single budget, breaker, jitter clock and cooldown timer.
        # Without this the resolver alone can exhaust the budget on slug
        # discovery before the observer ever runs.
        rate_guard = LinkedInRateGuard.from_settings()

        # --- Stage 4.5: LinkedIn slug backfill (BUG-A) ---
        # Seed workbook has no LinkedIn column, so the public observer
        # would silently no-op. Infer the slug from domain/name and
        # verify via a logged-out /company/<slug>/ probe.
        t0 = time.time()
        slug_http: HttpClient | None = None
        if args.mode in {"live-safe", "live-full"}:
            try:
                settings = get_settings()
                # Sink attached here too - slug probes are the single
                # biggest source of LinkedIn requests per run, and
                # archiving their HTML lets us re-run slug heuristics
                # offline if we ever tweak the disambiguator.
                from headcount.ingest.raw_response_store import (
                    build_sink_from_session,
                )

                slug_http = HttpClient(
                    cache=FileCache(settings.cache_dir),
                    configs=default_http_configs(),
                    transport=None,
                    raw_response_sink=build_sink_from_session(session),
                )
                slug_stats = backfill_linkedin_slugs(
                    session,
                    company_ids=company_ids,
                    http=slug_http,
                    rate_guard=rate_guard,
                )
                summary["stages"]["linkedin_slug_backfill"] = {
                    **slug_stats,
                    "elapsed_ms": int((time.time() - t0) * 1000),
                }
            except Exception as exc:
                ctx.add(
                    stage="linkedin_slug_backfill",
                    severity="warn",
                    message="LinkedIn slug backfill failed",
                    detail={
                        "error": str(exc),
                        "traceback": traceback.format_exc(limit=5),
                    },
                )
                summary["stages"]["linkedin_slug_backfill"] = {
                    "status": "error",
                    "error": str(exc),
                    "elapsed_ms": int((time.time() - t0) * 1000),
                }
        else:
            summary["stages"]["linkedin_slug_backfill"] = {
                "status": "skipped_offline_mode",
            }

        # --- Stage 5: collect anchors (live, scoped) ---
        t0 = time.time()
        try:
            anchor_summary = asyncio.run(
                _collect_anchors_scoped(
                    session,
                    companies,
                    mode=args.mode,
                    rate_guard=rate_guard,
                    run_label=run_label,
                )
            )
            summary["stages"]["collect_anchors"] = {
                **dict(anchor_summary),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        except Exception as exc:
            ctx.add(
                stage="collect_anchors",
                severity="error",
                message="collect_anchors raised",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )
            summary["stages"]["collect_anchors"] = {
                "status": "error",
                "error": str(exc),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }

        # --- Stage 5.5: optional breaker-recovery pass ---
        # If the LinkedIn breaker tripped during stage 5 we now have a
        # list of company ids parked in the guard. Behind a flag we
        # sleep through the cooldown and re-collect anchors only for
        # those companies so a 250-company production run does not
        # lose its long tail to a single 999 burst. Off by default so
        # interactive cohort runs stay snappy.
        deferred = rate_guard.drain_deferred()
        retry_meta: dict[str, Any] = {
            "deferred_count": len(deferred),
            "trips_observed": rate_guard.trip_count,
        }
        if deferred and args.retry_breaker_skips and args.mode == "live-full":
            # Floor the cooldown at a meaningful pause even when the
            # breaker never tripped (eg. some companies 999'd but the
            # streak stayed under the threshold). Without this the
            # retry would fire immediately and almost certainly 999
            # again against a soft-flagged IP.
            recovery_floor = max(
                120.0, float(args.breaker_recovery_floor_seconds or 0.0)
            )
            cooldown = max(recovery_floor, rate_guard.cooldown_remaining() + 5.0)
            retry_meta["cooldown_slept_seconds"] = cooldown
            print(
                f"[breaker-recovery] sleeping {cooldown:.1f}s for cooldown,"
                f" then retrying {len(deferred)} deferred companies",
                flush=True,
            )
            time.sleep(cooldown)
            # is_circuit_open() now re-arms the guard automatically.
            assert not rate_guard.is_circuit_open(), (
                "guard should be re-armed after sleeping cooldown"
            )

            # Some deferred companies tripped the breaker during the
            # slug backfill and therefore never had a linkedin URL
            # persisted to the DB. Re-run slug backfill for just those
            # IDs before the observer retry so the observer actually
            # has something to fetch.
            deferred_set = set(deferred)
            missing_slug_ids = [
                c.id
                for c in companies
                if c.id in deferred_set
                and not (
                    getattr(c, "linkedin_company_url", None)
                    and str(c.linkedin_company_url).strip()
                )
            ]
            # linkedin_company_url lives on the ORM row; re-fetch so
            # we see URLs that were persisted in the primary backfill
            # (those companies were deferred by the *observer* after
            # their slug was already known - we skip those here).
            if missing_slug_ids and slug_http is not None:
                try:
                    retry_slug_stats = backfill_linkedin_slugs(
                        session,
                        company_ids=missing_slug_ids,
                        http=slug_http,
                        rate_guard=rate_guard,
                    )
                    retry_meta["retry_slug_backfill"] = retry_slug_stats
                except Exception as exc:
                    ctx.add(
                        stage="collect_anchors_retry",
                        severity="warn",
                        message="breaker-recovery slug backfill raised",
                        detail={"error": str(exc)},
                    )
                    retry_meta["retry_slug_backfill_error"] = str(exc)

            # Refresh ORM state so the observer sees slugs that just
            # got persisted by the retry backfill.
            session.commit()
            retry_companies = [c for c in companies if c.id in set(deferred)]
            for c in retry_companies:
                session.refresh(c)

            t1 = time.time()
            try:
                retry_summary = asyncio.run(
                    _collect_anchors_scoped(
                        session,
                        retry_companies,
                        mode=args.mode,
                        rate_guard=rate_guard,
                        run_label=f"{run_label}:retry",
                    )
                )
                retry_meta["retry_summary"] = dict(retry_summary)
                retry_meta["status"] = "ok"
            except Exception as exc:
                ctx.add(
                    stage="collect_anchors_retry",
                    severity="warn",
                    message="breaker-recovery retry pass raised",
                    detail={"error": str(exc)},
                )
                retry_meta["status"] = "error"
                retry_meta["error"] = str(exc)
            retry_meta["elapsed_ms"] = int((time.time() - t1) * 1000)
        elif deferred:
            retry_meta["status"] = "skipped_no_flag"
            retry_meta["deferred_company_ids"] = deferred
        else:
            retry_meta["status"] = "no_deferred"
        summary["stages"]["collect_anchors_retry"] = retry_meta
        summary["stages"]["collect_anchors_retry"]["linkedin_requests_made"] = (
            rate_guard.requests_made
        )

        # --- Stage 6: collect employment (OCR only; no benchmark promotion
        #     because we haven't loaded any benchmark rows yet) ---
        t0 = time.time()
        ocr_observer = (
            _build_ocr_observer(ctx) if args.mode == "live-full" else None
        )
        emp_sources: list[str] = []
        if ocr_observer is not None:
            emp_sources.append("linkedin_ocr")
        try:
            # Per Phase 11 scope: Harmonic / Zeeshan / LinkedIn benchmark
            # workbook rows are evaluation-only. They must never be
            # promoted into anchor observations, or the scoreboard will
            # read back the very numbers we are supposed to be
            # approximating (see BUG-leak in HARMONIC_COHORT_LIVE_RUN).
            from headcount.db.enums import BenchmarkProvider as _BP
            emp_result = _collect_emp(
                session,
                company_ids=company_ids,
                sources=emp_sources,
                note=f"harmonic_cohort_live mode={args.mode} (blind)",
                ocr_observer=ocr_observer,
                benchmark_skip_providers={
                    _BP.harmonic,
                    _BP.zeeshan,
                    _BP.linkedin,
                },
            )
            summary["stages"]["collect_employment"] = {
                **emp_result.summary(),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        except Exception as exc:
            ctx.add(
                stage="collect_employment",
                severity="error",
                message="collect_employment raised",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )
            summary["stages"]["collect_employment"] = {
                "status": "error",
                "error": str(exc),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }

        # --- Stage 7: estimate series (scoped; uses only live-sourced anchors) ---
        t0 = time.time()
        try:
            est = run_estimate(
                session,
                start_month=start,
                end_month=as_of,
                as_of_month=as_of,
                company_ids=company_ids,
                sample_floor=5,
                note=f"harmonic_cohort_live mode={args.mode} (blind)",
            )
            summary["stages"]["estimate_series"] = {
                "run_id": est.run_id,
                "final_status": est.final_status.value,
                "companies_attempted": est.companies_attempted,
                "companies_succeeded": est.companies_succeeded,
                "companies_failed": est.companies_failed,
                "companies_degraded": est.companies_degraded,
                "months_written": est.months_written,
                "months_flagged": est.months_flagged,
                "review_items_inserted": est.review_items_inserted,
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        except Exception as exc:
            ctx.add(
                stage="estimate_series",
                severity="error",
                message="estimate_series raised",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )
            summary["stages"]["estimate_series"] = {
                "status": "error",
                "error": str(exc),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }

        # --- Stage 8: snapshot production coverage BEFORE loading benchmarks,
        #     so we're certain no benchmark rows contaminated the anchor set. ---
        t0 = time.time()
        try:
            production_coverage = _production_coverage(
                session, company_ids=company_ids
            )
            summary["stages"]["production_coverage"] = {
                "totals": production_coverage["totals"],
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        except Exception as exc:
            production_coverage = {"error": str(exc)}
            ctx.add(
                stage="production_coverage",
                severity="error",
                message="coverage snapshot failed",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )

        # --- Stage 9: NOW load benchmarks, evaluation-only ---
        t0 = time.time()
        bench = load_benchmarks(session, benchmarks_workbook)
        # Backfill benchmark_observation.company_id FKs via the existing
        # candidate resolver linkage. This is a plain FK fill - it does
        # NOT promote benchmark rows into the anchor table.
        _backfill_benchmark_links(session)
        summary["stages"]["load_benchmarks_for_eval"] = {
            "sheets_loaded": list(bench.sheets_loaded),
            "observations_written": bench.observations_written,
            "event_candidates": bench.event_candidates_written,
            "note": "benchmarks loaded AFTER estimation; FKs backfilled but not promoted to anchors",
            "elapsed_ms": int((time.time() - t0) * 1000),
        }

        # --- Stage 10: evaluate ---
        t0 = time.time()
        try:
            scoreboard = _score_harmonic_cohort(
                session, as_of=as_of, company_ids=company_ids
            )
            summary["stages"]["evaluate"] = {
                "status": "ok",
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        except Exception as exc:
            scoreboard = {"error": str(exc)}
            ctx.add(
                stage="evaluate",
                severity="error",
                message="evaluate_against_benchmarks raised",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )
            summary["stages"]["evaluate"] = {
                "status": "error",
                "error": str(exc),
                "elapsed_ms": int((time.time() - t0) * 1000),
            }

        # --- Stage 11: per-company delta report ---
        try:
            per_company = _per_company_report(
                session,
                harmonic_targets=harmonic_targets,
                as_of=as_of,
                company_map=company_map,
            )
        except Exception as exc:
            per_company = []
            ctx.add(
                stage="per_company",
                severity="error",
                message="per-company report failed",
                detail={
                    "error": str(exc),
                    "traceback": traceback.format_exc(limit=5),
                },
            )

    (run_dir / "pipeline.json").write_text(
        json.dumps(summary, indent=2, default=str), encoding="utf-8"
    )
    (run_dir / "scoreboard.json").write_text(
        json.dumps(scoreboard, indent=2, default=str), encoding="utf-8"
    )
    (run_dir / "per_company.json").write_text(
        json.dumps(per_company, indent=2, default=str), encoding="utf-8"
    )
    (run_dir / "production_coverage.json").write_text(
        json.dumps(production_coverage, indent=2, default=str), encoding="utf-8"
    )
    (run_dir / "issues.json").write_text(
        json.dumps([i.to_dict() for i in ctx.issues], indent=2, default=str),
        encoding="utf-8",
    )
    print(
        "harmonic_cohort_live "
        f"mode={args.mode} "
        f"run_dir={run_dir} "
        f"cohort={summary['stages'].get('pick_cohort', {}).get('resolved_companies')} "
        f"estimate_status="
        f"{summary['stages'].get('estimate_series', {}).get('final_status')} "
        f"issues={len(ctx.issues)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv[1:]))
