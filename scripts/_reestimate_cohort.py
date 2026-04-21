"""Re-run estimate_series + evaluation for the Harmonic cohort after
new anchors have been written to an existing cohort run directory.

Useful when a surgical slug retry writes new LinkedIn anchors but
doesn't itself re-estimate (eg because the narrower retry script
failed mid-flight)."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))


def _bootstrap(run_dir: Path) -> None:
    db_path = (run_dir / "cohort.sqlite").resolve()
    cache_dir = (run_dir / "http_cache").resolve()
    run_artifact_dir = (run_dir / "run_artifacts").resolve()
    for p in (cache_dir, run_artifact_dir):
        p.mkdir(parents=True, exist_ok=True)
    os.environ["DB_URL"] = f"sqlite:///{db_path.as_posix()}"
    os.environ["CACHE_DIR"] = str(cache_dir)
    os.environ["RUN_ARTIFACT_DIR"] = str(run_artifact_dir)


def _norm(s: str) -> str:
    out = s.lower().strip()
    for t in (",", ".", "(", ")", '"', "'"):
        out = out.replace(t, " ")
    for sfx in (
        " incorporated", " inc", " llc", " ltd", " limited", " corp",
        " corporation", " company", " co", " gmbh", " sa", " plc",
    ):
        if out.endswith(sfx):
            out = out[: -len(sfx)]
    return " ".join(out.split())


def _load_harmonic_names() -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(
        REPO_ROOT / "test_source" / "Sample Employee Growth for High Priority Prospects.xlsx",
        data_only=True, read_only=True,
    )
    ws = wb["Harmonic April 8"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = list(header).index("Company Name")
    return [str(r[idx]).strip() for r in rows[1:] if r[idx]]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("run_dir", type=Path)
    args = parser.parse_args()
    run_dir = args.run_dir.resolve()
    _bootstrap(run_dir)

    from headcount.db.engine import session_scope
    from headcount.estimate.pipeline import estimate_series
    from headcount.models.company import Company
    from headcount.models.company_alias import CompanyAlias
    from headcount.review.evaluation import evaluate_against_benchmarks

    names = _load_harmonic_names()
    by_norm = {_norm(n): n for n in names}

    with session_scope() as session:
        cohort: dict[str, object] = {}
        for c in session.query(Company).all():
            if _norm(c.canonical_name or "") in by_norm:
                cohort[c.id] = c
        for a in session.query(CompanyAlias).all():
            if _norm(a.alias_name or "") in by_norm and a.company_id not in cohort:
                c = session.get(Company, a.company_id)
                if c is not None:
                    cohort[c.id] = c
        cohort_ids = list(cohort.keys())
        print(f"[reestimate] cohort size: {len(cohort_ids)}")

        result = estimate_series(
            session,
            company_ids=cohort_ids,
            start_month=date(2022, 1, 1),
            end_month=date(2026, 4, 1),
            as_of_month=date(2026, 4, 1),
            note="post_retry_reestimate",
        )
        print(f"[reestimate] estimate result: {result!r}")
        session.commit()

        scoreboard = evaluate_against_benchmarks(
            session,
            as_of_month=date(2026, 4, 1),
            company_ids=cohort_ids,
        )
        sb = scoreboard.to_dict()
        out = run_dir / "retry_scoreboard.json"
        out.write_text(json.dumps(sb, default=str, indent=2), encoding="utf-8")
        print(f"[reestimate] wrote {out}")
        hl = sb.get("headline", {})
        comps = sb.get("companies", {})
        print(f"[reestimate] headline: {hl}")
        print(f"[reestimate] companies: {comps}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
