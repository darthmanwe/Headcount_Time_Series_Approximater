"""Targeted retry: try to resolve LinkedIn slugs for declined cohort companies.

After a cohort run finishes with some companies stuck at
``linkedin_company_url IS NULL`` because every slug probe returned 999
/ 202 / CAPTCHA, this script re-attempts slug discovery on just those
companies. It assumes stale DDG/Bing SERP envelopes were already
purged from the run's ``http_cache`` so the SERP fallback actually
hits the network again.

Steps:
1. Point the runtime at the given cohort run directory (DB + cache).
2. Read the Harmonic target list and match names to companies in DB.
3. Filter to the ones still missing ``linkedin_company_url``.
4. Call ``backfill_linkedin_slugs`` with a fresh guard.
5. For any newly-resolved slug, run ``collect_anchors`` scoped to
   just those companies (LinkedIn JSON-LD etc).
6. Re-run ``estimate_series`` for the affected companies.
7. Rebuild the scoreboard over the full Harmonic cohort and write
   ``retry_scoreboard.json`` + ``retry_per_company.json``.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))


def _bootstrap(run_dir: Path) -> None:
    db_path = (run_dir / "cohort.sqlite").resolve()
    cache_dir = (run_dir / "http_cache").resolve()
    run_artifact_dir = (run_dir / "run_artifacts").resolve()
    for p in (cache_dir, run_artifact_dir):
        p.mkdir(parents=True, exist_ok=True)
    os.environ["DB_URL"] = f"sqlite:///{db_path.as_posix()}"
    os.environ["CACHE_DIR"] = str(cache_dir)
    os.environ["RUN_ARTIFACT_DIR"] = str(run_artifact_dir)


def _norm(s: str) -> str:
    out = s.lower().strip()
    for token in (",", ".", "(", ")", '"', "'"):
        out = out.replace(token, " ")
    for suffix in (
        " incorporated", " inc", " llc", " ltd", " limited", " corp",
        " corporation", " company", " co", " gmbh", " sa", " plc",
    ):
        if out.endswith(suffix):
            out = out[: -len(suffix)]
    return " ".join(out.split())


def _load_harmonic_names() -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(
        REPO_ROOT / "test_source" / "Sample Employee Growth for High Priority Prospects.xlsx",
        data_only=True,
        read_only=True,
    )
    ws = wb["Harmonic April 8"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    name_idx = list(header).index("Company Name")
    return [
        str(r[name_idx]).strip()
        for r in rows[1:]
        if r[name_idx]
    ]


def _match_companies(session, names: list[str]) -> list[object]:
    from headcount.models.company import Company
    from headcount.models.company_alias import CompanyAlias

    by_norm = {_norm(n): n for n in names}
    comps = session.query(Company).all()
    hits: dict[str, object] = {}
    for c in comps:
        if _norm(c.canonical_name or "") in by_norm:
            hits[c.id] = c
    aliases = session.query(CompanyAlias).all()
    for a in aliases:
        if _norm(a.alias_name or "") in by_norm and a.company_id not in hits:
            c = session.query(Company).get(a.company_id)
            if c is not None:
                hits[c.id] = c
    return list(hits.values())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("run_dir", type=Path)
    parser.add_argument("--circuit-threshold", type=int, default=20)
    parser.add_argument("--budget", type=int, default=120)
    args = parser.parse_args()

    run_dir = args.run_dir.resolve()
    if not run_dir.exists():
        print(f"Run dir not found: {run_dir}", file=sys.stderr)
        return 2

    _bootstrap(run_dir)

    from headcount.db.engine import session_scope
    from headcount.ingest.collect import default_http_configs
    from headcount.ingest.http import FileCache, HttpClient
    from headcount.ingest.linkedin_guard import LinkedInRateGuard
    from headcount.resolution.linkedin_resolver import backfill_linkedin_slugs

    harmonic_names = _load_harmonic_names()
    print(f"[retry] harmonic targets: {len(harmonic_names)}")

    with session_scope() as session:
        cohort = _match_companies(session, harmonic_names)
        cohort_ids = [c.id for c in cohort]
        missing = [c for c in cohort if c.linkedin_company_url is None]
        missing_ids = [c.id for c in missing]
        print(f"[retry] harmonic cohort resolved: {len(cohort_ids)}")
        print(f"[retry] missing slugs in cohort: {len(missing_ids)}")
        for c in missing:
            print(f"  - {c.canonical_name} ({c.canonical_domain})")

        if not missing_ids:
            print("[retry] nothing to do")
            return 0

        guard = LinkedInRateGuard.from_settings(
            circuit_threshold=args.circuit_threshold,
            daily_request_budget=args.budget,
        )

        cache = FileCache(Path(os.environ["CACHE_DIR"]))
        http = HttpClient(cache=cache, configs=default_http_configs())
        stats = backfill_linkedin_slugs(
            session,
            company_ids=missing_ids,
            http=http,
            rate_guard=guard,
        )
        session.commit()
        print(f"[retry] slug backfill stats: {stats}")
        print(f"[retry] deferred companies: {len(guard.deferred_companies)}")

        from headcount.models.company import Company

        newly_slugged = [
            c
            for c in session.query(Company).filter(Company.id.in_(missing_ids)).all()
            if c.linkedin_company_url is not None
        ]
        print(f"[retry] newly slugged: {len(newly_slugged)}")
        for c in newly_slugged:
            print(f"  + {c.canonical_name}: {c.linkedin_company_url}")

        if not newly_slugged:
            print("[retry] no new slugs; skipping anchor collect")
            return 0

        from headcount.ingest.collect import collect_anchors
        from headcount.ingest.observers import (
            CompanyWebObserver,
            LinkedInPublicObserver,
            ManualAnchorObserver,
            SECObserver,
            WikidataObserver,
        )

        adapters = [
            ManualAnchorObserver(),
            SECObserver(),
            WikidataObserver(),
            CompanyWebObserver(),
            LinkedInPublicObserver(rate_guard=guard),
        ]

        async def _run_anchors():
            return await collect_anchors(
                session,
                adapters=adapters,
                companies=newly_slugged,
                http_client=http,
            )

        result = asyncio.run(_run_anchors())
        print(f"[retry] collect_anchors: {result.summary()}")

        from headcount.estimate.pipeline import estimate_series

        new_ids = [c.id for c in newly_slugged]
        est_summary = estimate_series(
            session,
            company_ids=new_ids,
            start_month=date(2022, 1, 1),
            end_month=date(2026, 4, 1),
            as_of_month=date(2026, 4, 1),
            note="retry_declined_slugs",
        )
        print(f"[retry] estimate_series: {est_summary!r}")
        session.commit()

        from headcount.review.evaluation import evaluate_against_benchmarks

        scoreboard = evaluate_against_benchmarks(
            session,
            as_of_month=date(2026, 4, 1),
            company_ids=cohort_ids,
        )
        out = run_dir / "retry_scoreboard.json"
        out.write_text(
            json.dumps(scoreboard.to_dict(), default=str, indent=2),
            encoding="utf-8",
        )
        print(f"[retry] wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
