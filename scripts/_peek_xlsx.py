"""Quick inspection helper for the test_source/ workbooks."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def peek(path: Path, rows_per_sheet: int = 6) -> None:
    print(f"== {path.name} ==")
    wb = load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"-- sheet: {sn!r} dims={ws.max_row}x{ws.max_column}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= rows_per_sheet:
                break
            print(f"  row {i}: {row}")
    print()


if __name__ == "__main__":
    for p in sys.argv[1:]:
        peek(Path(p))
