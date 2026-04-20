"""Load ``Sample Employee Growth for High Priority Prospects`` workbook.

The workbook carries three provider views: ``Zeeshan April 1``,
``Harmonic April 8``, and ``LinkedIn April 13``. Each sheet has its own
column layout; the loader hard-codes the mapping because the column
semantics (not just the headers) differ per sheet and we prefer a fail-loud
mismatch over a silent schema drift. The ``SUMMARY`` tab is ignored
because every value is duplicated elsewhere.

Provenance (workbook, sheet, row index, cell address, column header) is
preserved on every emitted row so analysts can audit any disagreement
between system output and benchmark without chasing spreadsheets.

Loader is idempotent on
``(source_workbook, source_sheet, source_row_index, provider, metric)``.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from headcount.db.enums import (
    BenchmarkMetric,
    BenchmarkProvider,
    HeadcountValueKind,
)
from headcount.models.benchmark import (
    BenchmarkEventCandidate,
    BenchmarkObservation,
)
from headcount.models.company_candidate import CompanyCandidate
from headcount.parsers.benchmark_notes import parse_note_hint
from headcount.parsers.headcount_value import ParsedHeadcount, parse_headcount_value
from headcount.utils.logging import get_logger
from headcount.utils.time import add_months, month_floor

_log = get_logger("headcount.ingest.seeds.benchmark_loader")


@dataclass(slots=True)
class BenchmarkLoadResult:
    workbook: str
    sheets_loaded: list[str] = field(default_factory=list)
    observations_written: int = 0
    observations_updated: int = 0
    observations_skipped: int = 0
    event_candidates_written: int = 0


@dataclass(frozen=True, slots=True)
class _SheetSpec:
    sheet: str
    provider: BenchmarkProvider
    as_of: date
    columns: dict[str, BenchmarkMetric]
    name_col: str
    domain_col: str | None = None
    linkedin_col: str | None = None
    range_col: str | None = None
    notes_col: str | None = None


_SHEETS: tuple[_SheetSpec, ...] = (
    _SheetSpec(
        sheet="Zeeshan April 1",
        provider=BenchmarkProvider.zeeshan,
        as_of=date(2026, 4, 1),
        name_col="Company name",
        domain_col="Company Domain Name",
        range_col="Current Employee Count",
        columns={
            "Employee Count (6 months ago)": BenchmarkMetric.headcount_6m_ago,
            "Employee Count (1 year ago)": BenchmarkMetric.headcount_1y_ago,
            "Employee Count (2 years ago)": BenchmarkMetric.headcount_2y_ago,
            "Employee Growth % (6 months)": BenchmarkMetric.growth_6m_pct,
            "Employee Growth % (1 year)": BenchmarkMetric.growth_1y_pct,
            "Employee Growth % (2 years)": BenchmarkMetric.growth_2y_pct,
        },
        notes_col="Assumptions",
    ),
    _SheetSpec(
        sheet="Harmonic April 8",
        provider=BenchmarkProvider.harmonic,
        as_of=date(2026, 4, 8),
        name_col="Company Name",
        columns={
            "Headcount": BenchmarkMetric.headcount_current,
            "Headcount % (365d)": BenchmarkMetric.growth_1y_pct,
            "Headcount % (180d)": BenchmarkMetric.growth_6m_pct,
            "Web Traffic": BenchmarkMetric.web_traffic,
        },
    ),
    _SheetSpec(
        sheet="LinkedIn April 13",
        provider=BenchmarkProvider.linkedin,
        as_of=date(2026, 4, 13),
        name_col="Company name",
        domain_col="Company Domain Name",
        linkedin_col="LinkedIn Domain",
        range_col="Employee Range",
        columns={
            "Employee Count": BenchmarkMetric.headcount_current,
            "Employee Count (6 months ago)": BenchmarkMetric.headcount_6m_ago,
            "Employee Count (1 year ago)": BenchmarkMetric.headcount_1y_ago,
            "Employee Count (2 years ago)": BenchmarkMetric.headcount_2y_ago,
            "Employee Growth % (6 months)": BenchmarkMetric.growth_6m_pct,
            "Employee Growth % (1 year)": BenchmarkMetric.growth_1y_pct,
            "Employee Growth % (2 years)": BenchmarkMetric.growth_2y_pct,
        },
        notes_col="Notes",
    ),
)


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_map(header: tuple[Any, ...]) -> dict[str, int]:
    return {_norm(cell): idx for idx, cell in enumerate(header) if cell is not None}


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _as_of_for_metric(as_of: date, metric: BenchmarkMetric) -> date:
    base = month_floor(as_of)
    if metric is BenchmarkMetric.headcount_6m_ago:
        return add_months(base, -6)
    if metric is BenchmarkMetric.headcount_1y_ago:
        return add_months(base, -12)
    if metric is BenchmarkMetric.headcount_2y_ago:
        return add_months(base, -24)
    return base


def _find_candidate(
    session: Session,
    *,
    name: str,
    domain: str | None,
) -> CompanyCandidate | None:
    name_norm = name.strip().lower()
    stmt = select(CompanyCandidate)
    if domain:
        domain_norm = domain.strip().lower()
        stmt_domain = stmt.where(CompanyCandidate.domain == domain_norm)
        candidate = session.execute(stmt_domain).scalars().first()
        if candidate is not None:
            return candidate
    for candidate in session.execute(stmt).scalars():
        if candidate.company_name.strip().lower() == name_norm:
            return candidate
    return None


def _upsert_observation(
    session: Session,
    *,
    workbook: str,
    sheet: str,
    row_index: int,
    provider: BenchmarkProvider,
    metric: BenchmarkMetric,
    **fields: Any,
) -> bool:
    """Return True if a new row was inserted, False if updated."""
    stmt = select(BenchmarkObservation).where(
        BenchmarkObservation.source_workbook == workbook,
        BenchmarkObservation.source_sheet == sheet,
        BenchmarkObservation.source_row_index == row_index,
        BenchmarkObservation.provider == provider,
        BenchmarkObservation.metric == metric,
    )
    existing = session.execute(stmt).scalar_one_or_none()
    if existing is None:
        session.add(
            BenchmarkObservation(
                source_workbook=workbook,
                source_sheet=sheet,
                source_row_index=row_index,
                provider=provider,
                metric=metric,
                **fields,
            )
        )
        return True
    for key, value in fields.items():
        if getattr(existing, key) != value:
            setattr(existing, key, value)
    return False


def _load_sheet(
    session: Session,
    workbook_path: Path,
    spec: _SheetSpec,
    *,
    result: BenchmarkLoadResult,
) -> None:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if spec.sheet not in wb.sheetnames:
        _log.warning("benchmark_sheet_missing", sheet=spec.sheet)
        return
    ws = wb[spec.sheet]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return

    header_map = _header_map(header)
    workbook = workbook_path.name

    def idx_of(label: str) -> int | None:
        return header_map.get(_norm(label))

    name_idx = idx_of(spec.name_col)
    if name_idx is None:
        raise ValueError(f"{workbook}:{spec.sheet} missing name column {spec.name_col!r}")
    domain_idx = idx_of(spec.domain_col) if spec.domain_col else None
    linkedin_idx = idx_of(spec.linkedin_col) if spec.linkedin_col else None
    range_idx = idx_of(spec.range_col) if spec.range_col else None
    notes_idx = idx_of(spec.notes_col) if spec.notes_col else None
    range_emitted_current = spec.range_col is not None and range_idx is not None

    for relative_index, row in enumerate(rows, start=1):
        raw_name = _get(row, name_idx)
        if not raw_name or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        raw_domain = _get(row, domain_idx)
        domain = (
            str(raw_domain).strip().lower()
            if raw_domain is not None and str(raw_domain).strip()
            else None
        )
        raw_linkedin = _get(row, linkedin_idx)
        linkedin_url = (
            str(raw_linkedin).strip()
            if raw_linkedin is not None and str(raw_linkedin).strip()
            else None
        )
        note_raw = _get(row, notes_idx)
        note = str(note_raw).strip() if note_raw is not None and str(note_raw).strip() else None

        candidate = _find_candidate(session, name=name, domain=domain)

        if range_idx is not None:
            raw_range = _get(row, range_idx)
            parsed_range = parse_headcount_value(raw_range)
            if parsed_range is not None:
                inserted = _upsert_observation(
                    session,
                    workbook=workbook,
                    sheet=spec.sheet,
                    row_index=relative_index,
                    provider=spec.provider,
                    metric=BenchmarkMetric.headcount_current,
                    company_candidate_id=candidate.id if candidate else None,
                    company_name_raw=name,
                    company_domain_raw=domain,
                    linkedin_url_raw=linkedin_url,
                    source_cell_address=f"{get_column_letter(range_idx + 1)}{relative_index + 1}",
                    source_column_name=spec.range_col,
                    as_of_month=spec.as_of,
                    value_min=parsed_range.value_min,
                    value_point=parsed_range.value_point,
                    value_max=parsed_range.value_max,
                    value_kind=parsed_range.kind,
                    raw_value_text=parsed_range.raw_value_text,
                    note=note,
                )
                result.observations_written += int(inserted)
                result.observations_updated += int(not inserted)

        for column_label, metric in spec.columns.items():
            col_idx = idx_of(column_label)
            if col_idx is None:
                continue
            raw_value = _get(row, col_idx)
            parsed: ParsedHeadcount | None = parse_headcount_value(raw_value)
            value_min: float | None
            value_point: float | None
            value_max: float | None
            value_kind: HeadcountValueKind | None
            raw_text: str | None
            if parsed is not None:
                value_min = parsed.value_min
                value_point = parsed.value_point
                value_max = parsed.value_max
                value_kind = parsed.kind
                raw_text = parsed.raw_value_text
            else:
                if raw_value is None:
                    continue
                raw_text = str(raw_value)
                try:
                    scalar = float(raw_text)
                except ValueError:
                    continue
                value_min = scalar
                value_point = scalar
                value_max = scalar
                value_kind = HeadcountValueKind.exact
            if metric is BenchmarkMetric.headcount_current and range_emitted_current:
                continue

            inserted = _upsert_observation(
                session,
                workbook=workbook,
                sheet=spec.sheet,
                row_index=relative_index,
                provider=spec.provider,
                metric=metric,
                company_candidate_id=candidate.id if candidate else None,
                company_name_raw=name,
                company_domain_raw=domain,
                linkedin_url_raw=linkedin_url,
                source_cell_address=f"{get_column_letter(col_idx + 1)}{relative_index + 1}",
                source_column_name=column_label,
                as_of_month=_as_of_for_metric(spec.as_of, metric),
                value_min=value_min,
                value_point=value_point,
                value_max=value_max,
                value_kind=value_kind,
                raw_value_text=raw_text,
                note=note if metric is BenchmarkMetric.headcount_current else None,
            )
            result.observations_written += int(inserted)
            result.observations_updated += int(not inserted)

        if note is not None:
            hint = parse_note_hint(note)
            if hint is not None:
                stmt = select(BenchmarkEventCandidate).where(
                    BenchmarkEventCandidate.source_workbook == workbook,
                    BenchmarkEventCandidate.source_sheet == spec.sheet,
                    BenchmarkEventCandidate.source_row_index == relative_index,
                    BenchmarkEventCandidate.description == hint.description,
                )
                if session.execute(stmt).scalar_one_or_none() is None:
                    session.add(
                        BenchmarkEventCandidate(
                            company_candidate_id=candidate.id if candidate else None,
                            source_workbook=workbook,
                            source_sheet=spec.sheet,
                            source_row_index=relative_index,
                            hint_type=hint.hint_type,
                            event_month_hint=hint.event_month_hint,
                            description=hint.description,
                        )
                    )
                    result.event_candidates_written += 1

    result.sheets_loaded.append(spec.sheet)


def load_benchmarks(
    session: Session,
    workbook_path: Path,
) -> BenchmarkLoadResult:
    """Load all supported sheets from the benchmark workbook."""
    result = BenchmarkLoadResult(workbook=workbook_path.name)
    for spec in _SHEETS:
        _load_sheet(session, workbook_path, spec, result=result)
    _log.info(
        "benchmark_load_summary",
        workbook=result.workbook,
        sheets=result.sheets_loaded,
        observations_written=result.observations_written,
        observations_updated=result.observations_updated,
        event_candidates=result.event_candidates_written,
    )
    return result
