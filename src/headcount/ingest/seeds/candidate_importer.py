"""Import ``High Priority Companies`` seed workbook into ``company_candidate``.

Idempotent on ``(source_workbook, source_sheet, source_row_index)`` so the
loader can be re-run after partial failures without creating duplicates.
Column mapping is detected by header name (lower-cased, whitespace
normalized) so small workbook edits don't silently skip rows.
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from headcount.models.company_candidate import CompanyCandidate
from headcount.utils.logging import get_logger

_log = get_logger("headcount.ingest.seeds.candidate_importer")

_NAME_ALIASES = {"company name", "company"}
_DOMAIN_ALIASES = {
    "company domain name",
    "company domain",
    "domain",
    "domain name",
}


@dataclass(slots=True)
class CandidateImportResult:
    workbook: str
    sheet: str
    rows_scanned: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    rows_updated: int = 0


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_map(header: Iterable[object]) -> tuple[int | None, int | None]:
    name_idx: int | None = None
    domain_idx: int | None = None
    for idx, cell in enumerate(header):
        key = _norm(cell)
        if name_idx is None and key in _NAME_ALIASES:
            name_idx = idx
        elif domain_idx is None and key in _DOMAIN_ALIASES:
            domain_idx = idx
    return name_idx, domain_idx


def import_candidates(
    session: Session,
    workbook_path: Path,
    *,
    sheet_name: str | None = None,
) -> CandidateImportResult:
    """Import rows from ``workbook_path`` into ``company_candidate``.

    Passing ``sheet_name`` forces a particular sheet; the default is the
    workbook's active sheet. Rows missing a company name are skipped.
    """
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    result = CandidateImportResult(workbook=workbook_path.name, sheet=ws.title)

    row_iter = ws.iter_rows(values_only=True)
    try:
        header = next(row_iter)
    except StopIteration:
        return result
    name_idx, domain_idx = _header_map(header)
    if name_idx is None:
        raise ValueError(
            f"{workbook_path.name}:{ws.title} missing a 'Company name' header; "
            f"found headers: {list(header)!r}"
        )

    for relative_index, row in enumerate(row_iter, start=1):
        result.rows_scanned += 1
        name = row[name_idx] if name_idx < len(row) else None
        if not name or not str(name).strip():
            result.rows_skipped += 1
            continue
        domain = None
        if domain_idx is not None and domain_idx < len(row):
            raw_domain = row[domain_idx]
            if raw_domain is not None and str(raw_domain).strip():
                domain = str(raw_domain).strip().lower()

        stmt = select(CompanyCandidate).where(
            CompanyCandidate.source_workbook == result.workbook,
            CompanyCandidate.source_sheet == result.sheet,
            CompanyCandidate.source_row_index == relative_index,
        )
        existing = session.execute(stmt).scalar_one_or_none()
        if existing is not None:
            if existing.company_name != str(name).strip() or existing.domain != domain:
                existing.company_name = str(name).strip()
                existing.domain = domain
                result.rows_updated += 1
            continue

        session.add(
            CompanyCandidate(
                source_workbook=result.workbook,
                source_sheet=result.sheet,
                source_row_index=relative_index,
                company_name=str(name).strip(),
                domain=domain,
            )
        )
        result.rows_imported += 1

    _log.info(
        "seed_candidate_import_summary",
        workbook=result.workbook,
        sheet=result.sheet,
        scanned=result.rows_scanned,
        imported=result.rows_imported,
        updated=result.rows_updated,
        skipped=result.rows_skipped,
    )
    return result
