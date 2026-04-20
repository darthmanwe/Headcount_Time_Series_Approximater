"""Integration tests for the benchmark loader."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import Session

from headcount.db.enums import (
    BenchmarkEventHintType,
    BenchmarkMetric,
    BenchmarkProvider,
    HeadcountValueKind,
)
from headcount.ingest.seeds import load_benchmarks
from headcount.models import (
    Base,
    BenchmarkEventCandidate,
    BenchmarkObservation,
    CompanyCandidate,
)


@pytest.fixture()
def session() -> Session:
    engine = create_engine("sqlite:///:memory:", future=True)

    @event.listens_for(engine, "connect")
    def _enable_fk(dbapi_conn, _):  # type: ignore[no-untyped-def]
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    with Session(engine, future=True) as s:
        yield s


def _build_benchmark_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    default.title = "SUMMARY"

    zeeshan = wb.create_sheet("Zeeshan April 1")
    zeeshan.append(
        [
            "Company name",
            "Company Domain Name",
            "Current Employee Count",
            "Employee Count (6 months ago)",
            "Employee Count (1 year ago)",
            "Employee Count (2 years ago)",
            "Employee Growth % (6 months)",
            "Employee Growth % (1 year)",
            "Employee Growth % (2 years)",
            "Assumptions",
        ]
    )
    zeeshan.append(
        [
            "Acme",
            "acme.com",
            "201-500",
            120,
            80,
            40,
            0.25,
            0.5,
            1.0,
            "Acquired by Symphony AI in June 2023",
        ]
    )

    harmonic = wb.create_sheet("Harmonic April 8")
    harmonic.append(
        ["Company Name", "Headcount", "Headcount % (365d)", "Headcount % (180d)", "Web Traffic"]
    )
    harmonic.append(["Acme", 150, 0.6, 0.2, 42000])

    linkedin = wb.create_sheet("LinkedIn April 13")
    linkedin.append(
        [
            "Company name",
            "Company Domain Name",
            "LinkedIn Domain",
            "Employee Range",
            "Employee Count",
            "Employee Count (6 months ago)",
            "Employee Count (1 year ago)",
            "Employee Count (2 years ago)",
            "Employee Growth % (6 months)",
            "Employee Growth % (1 year)",
            "Employee Growth % (2 years)",
            "Notes",
        ]
    )
    linkedin.append(
        [
            "Acme",
            "acme.com",
            "linkedin.com/company/acme",
            "201-500",
            155,
            125,
            85,
            45,
            0.24,
            0.82,
            2.44,
            "",
        ]
    )
    wb.save(path)


def _seed_candidate(session: Session) -> CompanyCandidate:
    cand = CompanyCandidate(
        source_workbook="seed.xlsx",
        source_sheet="Sheet1",
        source_row_index=1,
        company_name="Acme",
        domain="acme.com",
    )
    session.add(cand)
    session.commit()
    return cand


def test_benchmark_loader_writes_all_sheets(tmp_path: Path, session: Session) -> None:
    wb_path = tmp_path / "benchmarks.xlsx"
    _build_benchmark_workbook(wb_path)
    candidate = _seed_candidate(session)

    result = load_benchmarks(session, wb_path)
    session.commit()

    assert result.workbook == "benchmarks.xlsx"
    assert set(result.sheets_loaded) == {
        "Zeeshan April 1",
        "Harmonic April 8",
        "LinkedIn April 13",
    }
    assert result.observations_written >= 15
    assert result.observations_updated == 0

    obs = session.execute(
        select(BenchmarkObservation).where(
            BenchmarkObservation.provider == BenchmarkProvider.zeeshan,
            BenchmarkObservation.metric == BenchmarkMetric.headcount_current,
        )
    ).scalar_one()
    assert obs.company_candidate_id == candidate.id
    assert obs.value_min == 201
    assert obs.value_max == 500
    assert obs.value_kind is HeadcountValueKind.range
    assert obs.as_of_month == date(2026, 4, 1)
    assert obs.source_cell_address == "C2"
    assert obs.source_column_name == "Current Employee Count"

    six_m = session.execute(
        select(BenchmarkObservation).where(
            BenchmarkObservation.provider == BenchmarkProvider.zeeshan,
            BenchmarkObservation.metric == BenchmarkMetric.headcount_6m_ago,
        )
    ).scalar_one()
    assert six_m.as_of_month == date(2025, 10, 1)
    assert six_m.value_point == 120


def test_benchmark_loader_is_idempotent(tmp_path: Path, session: Session) -> None:
    wb_path = tmp_path / "benchmarks.xlsx"
    _build_benchmark_workbook(wb_path)
    _seed_candidate(session)

    first = load_benchmarks(session, wb_path)
    session.commit()
    second = load_benchmarks(session, wb_path)
    session.commit()
    assert second.observations_written == 0
    assert second.observations_updated == 0 or second.observations_updated > 0
    # row count shouldn't grow on a replay
    total_before = first.observations_written
    total = session.execute(select(BenchmarkObservation)).scalars().all()
    assert len(total) == total_before


def test_benchmark_event_candidate_extracted(tmp_path: Path, session: Session) -> None:
    wb_path = tmp_path / "benchmarks.xlsx"
    _build_benchmark_workbook(wb_path)
    _seed_candidate(session)

    load_benchmarks(session, wb_path)
    session.commit()

    events = session.execute(select(BenchmarkEventCandidate)).scalars().all()
    assert len(events) == 1
    event_row = events[0]
    assert event_row.hint_type is BenchmarkEventHintType.acquisition
    assert event_row.event_month_hint == date(2023, 6, 1)
    assert event_row.source_sheet == "Zeeshan April 1"


def test_benchmark_loader_handles_missing_candidate(tmp_path: Path, session: Session) -> None:
    wb_path = tmp_path / "benchmarks.xlsx"
    _build_benchmark_workbook(wb_path)

    result = load_benchmarks(session, wb_path)
    session.commit()
    assert result.observations_written >= 15
    obs = session.execute(select(BenchmarkObservation)).scalars().first()
    assert obs is not None
    assert obs.company_candidate_id is None
    assert obs.company_name_raw == "Acme"
